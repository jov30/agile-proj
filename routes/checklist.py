"""Back-of-house operations checklist blueprint.

Provides the admin-only dashboard for daily opening / closing checklists,
food temperature records, staff issue reporting, photo evidence gallery, and
compliance reporting. All routes live under ``/admin/checklist``.
"""

from __future__ import annotations

import os
import uuid
from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path

from flask import (
    Blueprint,
    Response,
    abort,
    current_app,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    url_for,
)
from sqlalchemy import func
from werkzeug.utils import secure_filename

from models import (
    ChecklistAuditLog,
    ChecklistPhoto,
    ChecklistSession,
    ChecklistTask,
    IssueReport,
    TemperatureReading,
    TemperatureSession,
    db,
)
from routes.auth import admin_required
from routes.checklist_data import (
    ALLOWED_PHOTO_EXTENSIONS,
    CHECKLISTS,
    ISSUE_CATEGORIES,
    MANAGERS,
    PHOTOS_REQUIRED,
    SECTION_DEADLINES,
    STAFF_ROSTER,
    TEMPERATURE_RECORDS,
)


checklist_bp = Blueprint("checklist", __name__, url_prefix="/admin/checklist")

_ROOT_DIR = Path(__file__).resolve().parent.parent
UPLOAD_DIR = _ROOT_DIR / "static" / "uploads" / "checklist"


# ── Helpers ────────────────────────────────────────────────────────────────────

def _ensure_upload_dir() -> Path:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    return UPLOAD_DIR


def _audit(action: str, record_type: str, record_id: int | None, user_name: str, details: str = "") -> None:
    entry = ChecklistAuditLog(
        action=action,
        record_type=record_type,
        record_id=record_id,
        user_name=user_name,
        details=details,
    )
    db.session.add(entry)


def _parse_date(value: str | None, default: date | None = None) -> date:
    if not value:
        return default or date.today()
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return default or date.today()


def _photo_url(filename: str) -> str:
    return url_for("checklist.photo", filename=filename)


def _completion(session: ChecklistSession) -> tuple[int, int, int]:
    total = len(session.tasks)
    done = sum(1 for task in session.tasks if task.done)
    pct = int(done / total * 100) if total else 0
    return done, total, pct


def _serialize_task(task: ChecklistTask) -> dict:
    return {
        "task_order": task.task_order,
        "task_name": task.task_name,
        "done": bool(task.done),
        "note": task.note or "",
    }


def _serialize_photo(photo: ChecklistPhoto) -> dict:
    return {
        "id": photo.id,
        "filename": photo.filename,
        "original_name": photo.original_name,
        "photo_number": photo.photo_number,
        "uploaded_at": photo.uploaded_at.strftime("%Y-%m-%d %H:%M"),
        "uploaded_by": photo.uploaded_by,
        "url": _photo_url(photo.filename),
    }


def _serialize_session(session: ChecklistSession) -> dict:
    done, total, pct = _completion(session)
    return {
        "id": session.id,
        "type": session.checklist_type,
        "type_title": CHECKLISTS.get(session.checklist_type, {}).get("title", session.checklist_type),
        "type_short": CHECKLISTS.get(session.checklist_type, {}).get("short", session.checklist_type),
        "accent": CHECKLISTS.get(session.checklist_type, {}).get("accent", "#ff6a13"),
        "section": session.section,
        "date": session.session_date.isoformat(),
        "day_of_week": session.day_of_week,
        "responsible": session.responsible,
        "submitted_by": session.submitted_by,
        "general_note": session.general_note,
        "is_late": bool(session.is_late),
        "submitted_at": session.submitted_at.strftime("%Y-%m-%d %H:%M:%S"),
        "submitted_time": session.submitted_at.strftime("%H:%M"),
        "verified": bool(session.verified),
        "verified_by": session.verified_by,
        "verified_at": session.verified_at.strftime("%Y-%m-%d %H:%M") if session.verified_at else None,
        "verified_time": session.verified_at.strftime("%H:%M") if session.verified_at else None,
        "overall_result": session.overall_result,
        "issues_found": session.issues_found,
        "action_responsible": session.action_responsible,
        "manager_notes": session.manager_notes,
        "done_count": done,
        "total_count": total,
        "completion_pct": pct,
        "photo_count": len(session.photos),
    }


def _serialize_temperature(session: TemperatureSession) -> dict:
    return {
        "id": session.id,
        "type": session.record_type,
        "type_title": TEMPERATURE_RECORDS.get(session.record_type, {}).get("title", session.record_type),
        "type_short": TEMPERATURE_RECORDS.get(session.record_type, {}).get("short", session.record_type),
        "accent": TEMPERATURE_RECORDS.get(session.record_type, {}).get("accent", "#eb5130"),
        "date": session.session_date.isoformat(),
        "recorded_by": session.recorded_by,
        "checked_by": session.checked_by,
        "submitted_at": session.submitted_at.strftime("%Y-%m-%d %H:%M:%S"),
        "submitted_time": session.submitted_at.strftime("%H:%M"),
        "notes": session.notes,
    }


def _is_section_late(section: str) -> bool:
    deadline = SECTION_DEADLINES.get(section)
    if not deadline:
        return False
    hour, minute = deadline
    now = datetime.now()
    deadline_dt = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
    return now > deadline_dt


# ── Dashboard ──────────────────────────────────────────────────────────────────

@checklist_bp.get("/")
@admin_required
def home():
    today = date.today()
    week_ago = today - timedelta(days=7)

    checklist_status: dict[str, dict] = {}
    for chk_type in CHECKLISTS:
        opening = (
            ChecklistSession.query.filter_by(
                checklist_type=chk_type, section="opening", session_date=today
            ).first()
        )
        closing = (
            ChecklistSession.query.filter_by(
                checklist_type=chk_type, section="closing", session_date=today
            ).first()
        )
        checklist_status[chk_type] = {
            "opening": _serialize_session(opening) if opening else None,
            "closing": _serialize_session(closing) if closing else None,
        }

    temperature_status: dict[str, dict | None] = {}
    for temp_type in TEMPERATURE_RECORDS:
        rec = TemperatureSession.query.filter_by(
            record_type=temp_type, session_date=today
        ).first()
        temperature_status[temp_type] = _serialize_temperature(rec) if rec else None

    pending_count = ChecklistSession.query.filter_by(verified=False).count()
    week_chk = ChecklistSession.query.filter(ChecklistSession.session_date >= week_ago).count()
    week_temp = TemperatureSession.query.filter(TemperatureSession.session_date >= week_ago).count()

    submitted_today = ChecklistSession.query.filter_by(session_date=today).count()
    temps_today = TemperatureSession.query.filter_by(session_date=today).count()

    alerts: list[dict[str, str]] = []
    expected_sections = len(CHECKLISTS) * 2
    if submitted_today < expected_sections:
        alerts.append(
            {
                "tone": "warning",
                "msg": f"{expected_sections - submitted_today} checklist section(s) not yet submitted today",
            }
        )
    if temps_today < len(TEMPERATURE_RECORDS):
        alerts.append(
            {
                "tone": "warning",
                "msg": f"{len(TEMPERATURE_RECORDS) - temps_today} temperature record(s) not yet submitted today",
            }
        )
    if pending_count:
        alerts.append(
            {"tone": "info", "msg": f"{pending_count} checklist(s) awaiting manager verification"}
        )

    recent_audit = [
        {
            "action": row.action.replace("_", " "),
            "user": row.user_name or "—",
            "details": row.details or "",
            "time": row.created_at.strftime("%H:%M"),
            "date": row.created_at.strftime("%d %b %Y"),
            "tone": "verify" if "VERIFY" in row.action else ("temp" if "TEMP" in row.action else "save"),
        }
        for row in ChecklistAuditLog.query.order_by(ChecklistAuditLog.created_at.desc()).limit(15).all()
    ]

    return render_template(
        "checklist/dashboard.html",
        checklists=CHECKLISTS,
        temperature_records=TEMPERATURE_RECORDS,
        checklist_status=checklist_status,
        temperature_status=temperature_status,
        pending_count=pending_count,
        week_chk=week_chk,
        week_temp=week_temp,
        alerts=alerts,
        recent_audit=recent_audit,
        today_iso=today.isoformat(),
        today_label=today.strftime("%A, %d %b %Y"),
    )


# ── Checklist form & save ──────────────────────────────────────────────────────

@checklist_bp.get("/sheet/<chk_type>")
@admin_required
def checklist_form(chk_type: str):
    if chk_type not in CHECKLISTS:
        return redirect(url_for("checklist.home"))
    chk_date = _parse_date(request.args.get("date"))
    section = request.args.get("section", "opening")
    if section not in {"opening", "closing"}:
        section = "opening"

    chk_data = CHECKLISTS[chk_type]
    tasks = chk_data[section]
    existing = ChecklistSession.query.filter_by(
        checklist_type=chk_type, section=section, session_date=chk_date
    ).first()
    existing_tasks = [_serialize_task(t) for t in existing.tasks] if existing else []
    existing_payload = _serialize_session(existing) if existing else None
    existing_photos = [_serialize_photo(p) for p in existing.photos] if existing else []
    deadline_hour, deadline_minute = SECTION_DEADLINES[section]

    return render_template(
        "checklist/sheet.html",
        chk_type=chk_type,
        chk_data=chk_data,
        section=section,
        chk_date=chk_date.isoformat(),
        day_name=chk_date.strftime("%A"),
        tasks=tasks,
        existing=existing_payload,
        existing_tasks=existing_tasks,
        existing_photos=existing_photos,
        staff=STAFF_ROSTER,
        managers=MANAGERS,
        photos_required=PHOTOS_REQUIRED,
        deadline_label=f"{deadline_hour % 12 or 12}:{deadline_minute:02d} {'AM' if deadline_hour < 12 else 'PM'}",
    )


@checklist_bp.post("/sheet/<chk_type>/save")
@admin_required
def checklist_save(chk_type: str):
    if chk_type not in CHECKLISTS:
        return redirect(url_for("checklist.home"))

    chk_date = _parse_date(request.form.get("date"))
    section = request.form.get("section", "opening")
    if section not in {"opening", "closing"}:
        section = "opening"

    responsible = request.form.get("responsible", "").strip()
    submitted_by = request.form.get("submitted_by", "").strip()
    general_done_by = request.form.get("general_done_by", "").strip()
    manager_submit = request.form.get("manager_submit", "").strip()
    general_note = request.form.get("general_note", "").strip()
    day_name = chk_date.strftime("%A")
    is_late = _is_section_late(section)

    tasks = CHECKLISTS[chk_type][section]

    session_row = ChecklistSession.query.filter_by(
        checklist_type=chk_type, section=section, session_date=chk_date
    ).first()
    if session_row is None:
        session_row = ChecklistSession(
            checklist_type=chk_type,
            section=section,
            session_date=chk_date,
        )
        db.session.add(session_row)

    session_row.day_of_week = day_name
    session_row.responsible = responsible or None
    session_row.submitted_by = submitted_by or None
    session_row.general_done_by = general_done_by or None
    session_row.manager_submit = manager_submit or None
    session_row.general_note = general_note or None
    session_row.is_late = bool(is_late)
    session_row.submitted_at = datetime.now()

    # Reset task list to the latest reference data and re-apply form input.
    for old_task in list(session_row.tasks):
        db.session.delete(old_task)
    for idx, task_name in enumerate(tasks):
        session_row.tasks.append(
            ChecklistTask(
                task_order=idx,
                task_name=task_name,
                done=bool(request.form.get(f"done_{idx}")),
                note=(request.form.get(f"note_{idx}", "").strip() or None),
            )
        )

    # Handle photo uploads. Any new file replaces the entire photo set so we
    # always have a coherent batch of evidence per submission.
    upload_dir = _ensure_upload_dir()
    incoming_photos: list[tuple[int, "FileStorage"]] = []  # type: ignore[name-defined]
    for idx in range(PHOTOS_REQUIRED + 6):
        file_obj = request.files.get(f"photo_{idx}")
        if file_obj and file_obj.filename and file_obj.filename.strip():
            ext = os.path.splitext(file_obj.filename)[1].lower().lstrip(".")
            if ext in ALLOWED_PHOTO_EXTENSIONS:
                incoming_photos.append((idx, file_obj))

    if incoming_photos:
        for old_photo in list(session_row.photos):
            try:
                (upload_dir / old_photo.filename).unlink(missing_ok=True)
            except OSError:
                pass
            db.session.delete(old_photo)

        db.session.flush()
        for idx, file_obj in incoming_photos:
            ext = os.path.splitext(file_obj.filename)[1].lower()
            session_id = session_row.id or "new"
            fname = f"{session_id}_{idx}_{uuid.uuid4().hex[:8]}{ext}"
            dest = upload_dir / fname
            file_obj.save(dest)
            session_row.photos.append(
                ChecklistPhoto(
                    filename=fname,
                    original_name=secure_filename(file_obj.filename),
                    photo_number=idx,
                    file_size=dest.stat().st_size,
                    uploaded_by=submitted_by or None,
                )
            )

    db.session.flush()
    _audit(
        "SAVE_CHECKLIST",
        "checklist",
        session_row.id,
        submitted_by or "—",
        f"{CHECKLISTS[chk_type]['title']} / {section} / {chk_date.isoformat()}",
    )
    db.session.commit()

    # Rename photos that were saved with the placeholder `new` prefix.
    for photo in session_row.photos:
        if photo.filename.startswith("new_"):
            new_name = photo.filename.replace("new_", f"{session_row.id}_", 1)
            try:
                (upload_dir / photo.filename).rename(upload_dir / new_name)
                photo.filename = new_name
            except OSError:
                pass
    db.session.commit()

    return redirect(url_for("checklist.session_view", session_id=session_row.id))


@checklist_bp.get("/sheet/view/<int:session_id>")
@admin_required
def session_view(session_id: int):
    session_row = ChecklistSession.query.get_or_404(session_id)
    chk_data = CHECKLISTS.get(session_row.checklist_type, {})
    deadline_hour, deadline_minute = SECTION_DEADLINES.get(session_row.section, (10, 30))
    return render_template(
        "checklist/session_view.html",
        session=_serialize_session(session_row),
        tasks=[_serialize_task(t) for t in session_row.tasks],
        photos=[_serialize_photo(p) for p in session_row.photos],
        chk_data=chk_data,
        chk_type=session_row.checklist_type,
        staff=STAFF_ROSTER,
        deadline_label=f"{deadline_hour % 12 or 12}:{deadline_minute:02d} {'AM' if deadline_hour < 12 else 'PM'}",
    )


@checklist_bp.post("/sheet/verify/<int:session_id>")
@admin_required
def session_verify(session_id: int):
    session_row = ChecklistSession.query.get_or_404(session_id)
    verified_by = request.form.get("verified_by", "").strip()
    overall_result = request.form.get("overall_result", "").strip()
    issues_found = request.form.get("issues_found", "").strip()
    action_responsible = request.form.get("action_responsible", "").strip()
    manager_notes = request.form.get("manager_notes", "").strip()

    session_row.verified = True
    session_row.verified_by = verified_by or None
    session_row.verified_at = datetime.now()
    session_row.overall_result = overall_result or None
    session_row.issues_found = issues_found or None
    session_row.action_responsible = action_responsible or None
    session_row.manager_notes = manager_notes or None

    _audit(
        "VERIFY",
        "checklist",
        session_row.id,
        verified_by or "—",
        f"Result: {overall_result or '—'}",
    )
    db.session.commit()
    return redirect(url_for("checklist.session_view", session_id=session_id))


# ── Temperature ────────────────────────────────────────────────────────────────

@checklist_bp.get("/temperature/<temp_type>")
@admin_required
def temperature_form(temp_type: str):
    if temp_type not in TEMPERATURE_RECORDS:
        return redirect(url_for("checklist.home"))
    temp_date = _parse_date(request.args.get("date"))
    temp_data = TEMPERATURE_RECORDS[temp_type]
    existing = TemperatureSession.query.filter_by(
        record_type=temp_type, session_date=temp_date
    ).first()
    readings = []
    if existing:
        readings = [
            {
                "food_order": r.food_order,
                "food_name": r.food_name,
                "c1_time": r.c1_time or "",
                "c1_temp": r.c1_temp,
                "c2_time": r.c2_time or "",
                "c2_temp": r.c2_temp,
                "c3_time": r.c3_time or "",
                "c3_temp": r.c3_temp,
                "c4_time": r.c4_time or "",
                "c4_temp": r.c4_temp,
                "c5_time": r.c5_time or "",
                "c5_temp": r.c5_temp,
                "discarded": r.discarded,
            }
            for r in existing.readings
        ]
    return render_template(
        "checklist/temperature.html",
        temp_type=temp_type,
        temp_data=temp_data,
        temp_date=temp_date.isoformat(),
        existing=_serialize_temperature(existing) if existing else None,
        readings=readings,
        staff=STAFF_ROSTER,
    )


@checklist_bp.post("/temperature/<temp_type>/save")
@admin_required
def temperature_save(temp_type: str):
    if temp_type not in TEMPERATURE_RECORDS:
        return redirect(url_for("checklist.home"))

    temp_date = _parse_date(request.form.get("date"))
    recorded_by = request.form.get("recorded_by", "").strip()
    checked_by = request.form.get("checked_by", "").strip()
    notes = request.form.get("notes", "").strip()

    foods = TEMPERATURE_RECORDS[temp_type]["foods"]

    def _to_float(raw: str | None):
        if raw is None or not raw.strip():
            return None
        try:
            return float(raw)
        except ValueError:
            return None

    session_row = TemperatureSession.query.filter_by(
        record_type=temp_type, session_date=temp_date
    ).first()
    if session_row is None:
        session_row = TemperatureSession(record_type=temp_type, session_date=temp_date)
        db.session.add(session_row)

    session_row.recorded_by = recorded_by or None
    session_row.checked_by = checked_by or None
    session_row.notes = notes or None
    session_row.submitted_at = datetime.now()

    for reading in list(session_row.readings):
        db.session.delete(reading)

    for idx, food_name in enumerate(foods):
        session_row.readings.append(
            TemperatureReading(
                food_order=idx,
                food_name=food_name,
                c1_time=request.form.get(f"c1_time_{idx}", "").strip() or None,
                c1_temp=_to_float(request.form.get(f"c1_temp_{idx}")),
                c2_time=request.form.get(f"c2_time_{idx}", "").strip() or None,
                c2_temp=_to_float(request.form.get(f"c2_temp_{idx}")),
                c3_time=request.form.get(f"c3_time_{idx}", "").strip() or None,
                c3_temp=_to_float(request.form.get(f"c3_temp_{idx}")),
                c4_time=request.form.get(f"c4_time_{idx}", "").strip() or None,
                c4_temp=_to_float(request.form.get(f"c4_temp_{idx}")),
                c5_time=request.form.get(f"c5_time_{idx}", "").strip() or None,
                c5_temp=_to_float(request.form.get(f"c5_temp_{idx}")),
                discarded=request.form.get(f"discarded_{idx}", "N") or "N",
            )
        )

    db.session.flush()
    _audit(
        "SAVE_TEMP",
        "temperature",
        session_row.id,
        recorded_by or "—",
        f"{TEMPERATURE_RECORDS[temp_type]['title']} / {temp_date.isoformat()}",
    )
    db.session.commit()
    return redirect(url_for("checklist.temperature_view", session_id=session_row.id))


@checklist_bp.get("/temperature/view/<int:session_id>")
@admin_required
def temperature_view(session_id: int):
    session_row = TemperatureSession.query.get_or_404(session_id)
    temp_data = TEMPERATURE_RECORDS.get(session_row.record_type, {})
    readings = [
        {
            "food_order": r.food_order,
            "food_name": r.food_name,
            "checks": [
                {"time": getattr(r, f"c{n}_time"), "temp": getattr(r, f"c{n}_temp")}
                for n in range(1, 6)
            ],
            "discarded": r.discarded,
        }
        for r in session_row.readings
    ]
    return render_template(
        "checklist/temperature_view.html",
        session=_serialize_temperature(session_row),
        readings=readings,
        temp_data=temp_data,
    )


# ── History ────────────────────────────────────────────────────────────────────

@checklist_bp.get("/history")
@admin_required
def history():
    date_from = _parse_date(request.args.get("date_from"), default=date.today() - timedelta(days=30))
    date_to = _parse_date(request.args.get("date_to"))
    rec_type = request.args.get("type", "all")
    staff_filter = request.args.get("staff", "").strip()

    chk_query = ChecklistSession.query.filter(
        ChecklistSession.session_date.between(date_from, date_to)
    )
    if staff_filter:
        chk_query = chk_query.filter(
            (ChecklistSession.submitted_by == staff_filter)
            | (ChecklistSession.responsible == staff_filter)
        )
    chk_query = chk_query.order_by(
        ChecklistSession.session_date.desc(),
        ChecklistSession.checklist_type,
        ChecklistSession.section,
    )
    chk_records = (
        [_serialize_session(row) for row in chk_query.all()]
        if rec_type in {"all", "checklist"}
        else []
    )

    temp_query = TemperatureSession.query.filter(
        TemperatureSession.session_date.between(date_from, date_to)
    )
    if staff_filter:
        temp_query = temp_query.filter(
            (TemperatureSession.recorded_by == staff_filter)
            | (TemperatureSession.checked_by == staff_filter)
        )
    temp_query = temp_query.order_by(TemperatureSession.session_date.desc(), TemperatureSession.record_type)
    temp_records = (
        [_serialize_temperature(row) for row in temp_query.all()]
        if rec_type in {"all", "temperature"}
        else []
    )

    return render_template(
        "checklist/history.html",
        chk_records=chk_records,
        temp_records=temp_records,
        date_from=date_from.isoformat(),
        date_to=date_to.isoformat(),
        rec_type=rec_type,
        staff_filter=staff_filter,
        staff=STAFF_ROSTER,
    )


# ── Manager panel ──────────────────────────────────────────────────────────────

@checklist_bp.get("/manager")
@admin_required
def manager_panel():
    today = date.today()
    pending = (
        ChecklistSession.query.filter_by(verified=False)
        .order_by(
            ChecklistSession.session_date.desc(),
            ChecklistSession.checklist_type,
            ChecklistSession.section,
        )
        .all()
    )
    issues = (
        ChecklistSession.query.filter_by(overall_result="issues_found")
        .order_by(ChecklistSession.session_date.desc())
        .limit(20)
        .all()
    )
    verified_today = ChecklistSession.query.filter_by(verified=True, session_date=today).count()
    recent_log = (
        ChecklistAuditLog.query.order_by(ChecklistAuditLog.created_at.desc()).limit(20).all()
    )
    return render_template(
        "checklist/manager.html",
        pending=[_serialize_session(s) for s in pending],
        issues=[_serialize_session(s) for s in issues],
        verified_today=verified_today,
        recent_log=[
            {
                "action": row.action.replace("_", " "),
                "user": row.user_name or "—",
                "details": row.details or "",
                "time": row.created_at.strftime("%H:%M"),
                "date": row.created_at.strftime("%d %b %Y"),
                "tone": "verify" if "VERIFY" in row.action else ("temp" if "TEMP" in row.action else "save"),
            }
            for row in recent_log
        ],
        staff=STAFF_ROSTER,
    )


# ── Photo gallery & file serving ───────────────────────────────────────────────

@checklist_bp.get("/photos")
@admin_required
def photo_gallery():
    date_filter = _parse_date(request.args.get("date"))
    chk_type = request.args.get("type", "all")
    query = (
        db.session.query(ChecklistPhoto, ChecklistSession)
        .join(ChecklistSession, ChecklistPhoto.session_id == ChecklistSession.id)
        .filter(ChecklistSession.session_date == date_filter)
    )
    if chk_type != "all":
        query = query.filter(ChecklistSession.checklist_type == chk_type)
    rows = query.order_by(ChecklistPhoto.session_id, ChecklistPhoto.photo_number).all()

    sessions: dict[int, dict] = {}
    for photo, sess in rows:
        bucket = sessions.setdefault(
            sess.id,
            {"session": _serialize_session(sess), "photos": []},
        )
        bucket["photos"].append(_serialize_photo(photo))

    dates_with_photos = [
        row[0].isoformat()
        for row in db.session.query(ChecklistSession.session_date)
        .join(ChecklistPhoto, ChecklistPhoto.session_id == ChecklistSession.id)
        .group_by(ChecklistSession.session_date)
        .order_by(ChecklistSession.session_date.desc())
        .limit(30)
        .all()
    ]

    return render_template(
        "checklist/photos.html",
        sessions=list(sessions.values()),
        date_filter=date_filter.isoformat(),
        chk_type=chk_type,
        checklists=CHECKLISTS,
        dates_with_photos=dates_with_photos,
    )


@checklist_bp.get("/photo/<path:filename>")
@admin_required
def photo(filename: str):
    safe = os.path.basename(filename)
    return send_from_directory(_ensure_upload_dir(), safe)


# ── Compliance / analytics ─────────────────────────────────────────────────────

@checklist_bp.get("/compliance")
@admin_required
def compliance_report():
    try:
        days = int(request.args.get("days", 30))
    except (TypeError, ValueError):
        days = 30
    days = max(1, min(days, 365))
    start = date.today() - timedelta(days=days - 1)
    end = date.today()

    per_staff_rows = (
        db.session.query(
            ChecklistSession.submitted_by.label("name"),
            func.count(ChecklistSession.id).label("total"),
            func.sum(func.cast(ChecklistSession.verified, db.Integer)).label("verified"),
            func.sum(
                func.cast(ChecklistSession.overall_result == "issues_found", db.Integer)
            ).label("issues"),
            func.max(ChecklistSession.submitted_at).label("last_submission"),
        )
        .filter(ChecklistSession.session_date.between(start, end))
        .filter(ChecklistSession.submitted_by.isnot(None))
        .filter(ChecklistSession.submitted_by != "")
        .group_by(ChecklistSession.submitted_by)
        .order_by(func.count(ChecklistSession.id).desc())
        .all()
    )

    per_type_rows = (
        db.session.query(
            ChecklistSession.checklist_type.label("type"),
            func.sum(
                func.cast(ChecklistSession.section == "opening", db.Integer)
            ).label("openings"),
            func.sum(
                func.cast(ChecklistSession.section == "closing", db.Integer)
            ).label("closings"),
            func.sum(func.cast(ChecklistSession.verified, db.Integer)).label("verified"),
            func.count(ChecklistSession.id).label("total"),
        )
        .filter(ChecklistSession.session_date.between(start, end))
        .group_by(ChecklistSession.checklist_type)
        .all()
    )

    photo_stats_rows = (
        db.session.query(
            ChecklistSession.submitted_by.label("submitted_by"),
            func.count(func.distinct(ChecklistSession.id)).label("sessions_with_photos"),
            func.count(ChecklistPhoto.id).label("total_photos"),
        )
        .outerjoin(ChecklistPhoto, ChecklistPhoto.session_id == ChecklistSession.id)
        .filter(ChecklistSession.session_date.between(start, end))
        .filter(ChecklistSession.submitted_by.isnot(None))
        .filter(ChecklistSession.submitted_by != "")
        .group_by(ChecklistSession.submitted_by)
        .order_by(func.count(ChecklistPhoto.id).desc())
        .all()
    )

    return render_template(
        "checklist/compliance.html",
        per_staff=[
            {
                "name": row.name,
                "total": row.total or 0,
                "verified": row.verified or 0,
                "issues": row.issues or 0,
                "last_submission": row.last_submission.strftime("%Y-%m-%d %H:%M")
                if row.last_submission
                else "—",
            }
            for row in per_staff_rows
        ],
        per_type=[
            {
                "type": row.type,
                "title": CHECKLISTS.get(row.type, {}).get("title", row.type),
                "accent": CHECKLISTS.get(row.type, {}).get("accent", "#ff6a13"),
                "openings": row.openings or 0,
                "closings": row.closings or 0,
                "verified": row.verified or 0,
                "total": row.total or 0,
            }
            for row in per_type_rows
        ],
        photo_stats=[
            {
                "submitted_by": row.submitted_by,
                "sessions_with_photos": row.sessions_with_photos or 0,
                "total_photos": row.total_photos or 0,
            }
            for row in photo_stats_rows
        ],
        days=days,
        start=start.isoformat(),
        end=end.isoformat(),
        checklists=CHECKLISTS,
    )


@checklist_bp.get("/analytics")
@admin_required
def analytics():
    return render_template("checklist/analytics.html", checklists=CHECKLISTS)


@checklist_bp.get("/api/analytics-data")
@admin_required
def analytics_data():
    try:
        days = int(request.args.get("days", 30))
    except (TypeError, ValueError):
        days = 30
    days = max(1, min(days, 365))
    start = date.today() - timedelta(days=days - 1)
    end = date.today()

    daily_chk = (
        db.session.query(
            ChecklistSession.session_date,
            func.count(ChecklistSession.id),
            func.sum(func.cast(ChecklistSession.verified, db.Integer)),
        )
        .filter(ChecklistSession.session_date.between(start, end))
        .group_by(ChecklistSession.session_date)
        .order_by(ChecklistSession.session_date)
        .all()
    )
    daily_temp = (
        db.session.query(TemperatureSession.session_date, func.count(TemperatureSession.id))
        .filter(TemperatureSession.session_date.between(start, end))
        .group_by(TemperatureSession.session_date)
        .order_by(TemperatureSession.session_date)
        .all()
    )
    staff_act = (
        db.session.query(
            ChecklistSession.submitted_by, func.count(ChecklistSession.id)
        )
        .filter(ChecklistSession.session_date.between(start, end))
        .filter(ChecklistSession.submitted_by.isnot(None))
        .filter(ChecklistSession.submitted_by != "")
        .group_by(ChecklistSession.submitted_by)
        .order_by(func.count(ChecklistSession.id).desc())
        .limit(13)
        .all()
    )
    type_comp = (
        db.session.query(
            ChecklistSession.checklist_type,
            func.count(ChecklistSession.id),
            func.sum(func.cast(ChecklistSession.verified, db.Integer)),
        )
        .filter(ChecklistSession.session_date.between(start, end))
        .group_by(ChecklistSession.checklist_type)
        .all()
    )

    labels: list[str] = []
    cursor = start
    while cursor <= end:
        labels.append(cursor.isoformat())
        cursor += timedelta(days=1)

    chk_map = {row[0].isoformat(): {"total": row[1] or 0, "verified": row[2] or 0} for row in daily_chk}
    temp_map = {row[0].isoformat(): row[1] or 0 for row in daily_temp}

    return jsonify(
        {
            "labels": labels,
            "daily_chk": [chk_map.get(d, {}).get("total", 0) for d in labels],
            "daily_verified": [chk_map.get(d, {}).get("verified", 0) for d in labels],
            "daily_temp": [temp_map.get(d, 0) for d in labels],
            "staff_names": [row[0] for row in staff_act],
            "staff_counts": [row[1] for row in staff_act],
            "type_labels": [
                CHECKLISTS.get(row[0], {}).get("short", row[0]) for row in type_comp
            ],
            "type_totals": [row[1] for row in type_comp],
            "type_verified": [row[2] or 0 for row in type_comp],
        }
    )


# ── Today timeline (admin dashboard widget) ────────────────────────────────────

@checklist_bp.get("/api/timeline")
@admin_required
def timeline_api():
    target = _parse_date(request.args.get("date"))

    checklist_events = (
        ChecklistSession.query.filter_by(session_date=target)
        .order_by(ChecklistSession.submitted_at)
        .all()
    )
    temperature_events = (
        TemperatureSession.query.filter_by(session_date=target)
        .order_by(TemperatureSession.submitted_at)
        .all()
    )

    events = []
    for row in checklist_events:
        done, total, pct = _completion(row)
        events.append(
            {
                "kind": "checklist",
                "id": row.id,
                "type": row.checklist_type,
                "type_label": CHECKLISTS.get(row.checklist_type, {}).get("short", row.checklist_type),
                "accent": CHECKLISTS.get(row.checklist_type, {}).get("accent", "#ff6a13"),
                "section": row.section,
                "time": row.submitted_at.strftime("%H:%M") if row.submitted_at else "—",
                "by": row.submitted_by or row.responsible or "—",
                "done_pct": pct,
                "done_count": done,
                "total_count": total,
                "photo_count": len(row.photos),
                "verified": bool(row.verified),
                "verified_by": row.verified_by or "",
                "verified_at": row.verified_at.strftime("%H:%M") if row.verified_at else "",
            }
        )
    for row in temperature_events:
        events.append(
            {
                "kind": "temperature",
                "id": row.id,
                "type": row.record_type,
                "type_label": TEMPERATURE_RECORDS.get(row.record_type, {}).get("short", row.record_type),
                "accent": TEMPERATURE_RECORDS.get(row.record_type, {}).get("accent", "#eb5130"),
                "section": "temp",
                "time": row.submitted_at.strftime("%H:%M") if row.submitted_at else "—",
                "by": row.recorded_by or "—",
                "done_pct": 100,
                "photo_count": 0,
                "verified": False,
            }
        )

    return jsonify({"date": target.isoformat(), "events": events})


# ── Issue reports ──────────────────────────────────────────────────────────────

@checklist_bp.route("/report", methods=["GET", "POST"])
@admin_required
def report_issue():
    if request.method == "POST":
        category = request.form.get("category", "").strip()
        title = request.form.get("title", "").strip()
        description = request.form.get("description", "").strip()
        reported_by = request.form.get("reported_by", "").strip()
        priority = request.form.get("priority", "normal").strip()

        if category and title and description and reported_by:
            photo_fname = None
            file_obj = request.files.get("photo")
            if file_obj and file_obj.filename and file_obj.filename.strip():
                ext = os.path.splitext(file_obj.filename)[1].lower().lstrip(".")
                if ext in ALLOWED_PHOTO_EXTENSIONS:
                    photo_fname = f"issue_{uuid.uuid4().hex[:12]}.{ext}"
                    file_obj.save(_ensure_upload_dir() / photo_fname)

            report = IssueReport(
                category=category,
                title=title,
                description=description,
                reported_by=reported_by,
                priority=priority if priority in {"normal", "urgent", "low"} else "normal",
                photo=photo_fname,
            )
            db.session.add(report)
            db.session.commit()
            return redirect(url_for("checklist.report_issue", submitted=1))

    submitted = request.args.get("submitted")
    return render_template(
        "checklist/report_issue.html",
        staff=STAFF_ROSTER,
        issue_categories=ISSUE_CATEGORIES,
        submitted=submitted,
    )


@checklist_bp.get("/reports")
@admin_required
def reports():
    status_filter = request.args.get("status", "open")
    category_filter = request.args.get("category", "").strip()

    query = IssueReport.query
    if status_filter != "all":
        query = query.filter_by(status=status_filter)
    if category_filter:
        query = query.filter_by(category=category_filter)
    rows = query.order_by(IssueReport.submitted_at.desc()).all()

    status_counts_rows = (
        db.session.query(IssueReport.status, func.count(IssueReport.id))
        .group_by(IssueReport.status)
        .all()
    )
    counts = defaultdict(int)
    for status, count in status_counts_rows:
        counts[status] = count

    return render_template(
        "checklist/reports.html",
        reports=[
            {
                "id": r.id,
                "category": r.category,
                "category_label": ISSUE_CATEGORIES.get(r.category, {}).get("label", r.category),
                "category_accent": ISSUE_CATEGORIES.get(r.category, {}).get("accent", "#6c584f"),
                "title": r.title,
                "description": r.description,
                "reported_by": r.reported_by,
                "priority": r.priority,
                "status": r.status,
                "admin_notes": r.admin_notes,
                "resolved_by": r.resolved_by,
                "resolved_at": r.resolved_at.strftime("%Y-%m-%d %H:%M") if r.resolved_at else None,
                "submitted_at": r.submitted_at.strftime("%Y-%m-%d %H:%M"),
                "photo_url": _photo_url(r.photo) if r.photo else None,
            }
            for r in rows
        ],
        status_filter=status_filter,
        category_filter=category_filter,
        counts=dict(counts),
        issue_categories=ISSUE_CATEGORIES,
    )


@checklist_bp.post("/reports/<int:report_id>/update")
@admin_required
def update_report(report_id: int):
    report = IssueReport.query.get_or_404(report_id)
    status = request.form.get("status", "open")
    admin_notes = request.form.get("admin_notes", "").strip()
    resolved_by = request.form.get("resolved_by", "").strip()
    report.status = status if status in {"open", "in_progress", "resolved", "closed"} else "open"
    report.admin_notes = admin_notes or None
    report.resolved_by = resolved_by or None
    report.resolved_at = datetime.now() if report.status == "resolved" else None
    db.session.commit()
    return redirect(url_for("checklist.reports"))


# ── Excel exports ──────────────────────────────────────────────────────────────

def _try_openpyxl():
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore

        return Workbook, Font, PatternFill, Alignment, get_column_letter
    except ImportError:
        return None


@checklist_bp.get("/export/checklist/<int:session_id>.xlsx")
@admin_required
def export_checklist_excel(session_id: int):
    deps = _try_openpyxl()
    if deps is None:
        abort(503)
    Workbook, Font, PatternFill, Alignment, _get_column_letter = deps

    session_row = ChecklistSession.query.get_or_404(session_id)
    chk_data = CHECKLISTS.get(session_row.checklist_type, {})
    done, total, pct = _completion(session_row)

    wb = Workbook()
    ws = wb.active
    ws.title = "Checklist"
    ws["A1"] = "MCQ VIETNAMESE STREET FOOD"
    ws["A1"].font = Font(bold=True, size=18, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="2C201B", end_color="2C201B", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 30

    ws["A2"] = f"{chk_data.get('title', '').upper()}  —  {session_row.section.upper()} CHECKLIST"
    ws["A2"].font = Font(bold=True, size=13, color="EB5130")
    ws["A2"].fill = PatternFill(start_color="FFF1E1", end_color="FFF1E1", fill_type="solid")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:E2")

    def add_info_row(label: str, value: str) -> None:
        ws.append(["", label, value])
        row = ws.max_row
        ws.cell(row, 2).font = Font(bold=True, color="6C584F")
        ws.cell(row, 3).font = Font(color="231814")
        ws.merge_cells(f"C{row}:E{row}")

    ws.append([""])
    add_info_row("Date:", f"{session_row.session_date.isoformat()} ({session_row.day_of_week or ''})")
    add_info_row("Section:", session_row.section.title())
    add_info_row("Submitted at:", session_row.submitted_at.strftime("%Y-%m-%d %H:%M") if session_row.submitted_at else "—")
    add_info_row("Responsible:", session_row.responsible or "—")
    add_info_row("Submitted by:", session_row.submitted_by or "—")
    if session_row.general_note:
        add_info_row("General Note:", session_row.general_note)
    add_info_row("Completion:", f"{done}/{total} tasks done ({pct}%)")
    if session_row.is_late:
        add_info_row("⚠ Submission:", "LATE — Submitted after deadline")

    ws.append([""])
    headers = ["#", "Task Description", "Status", "Note"]
    ws.append(headers)
    hr = ws.max_row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=col)
        cell.value = h
        cell.fill = PatternFill(start_color="2C201B", end_color="2C201B", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for task in session_row.tasks:
        ws.append([
            task.task_order + 1,
            task.task_name,
            "Done" if task.done else "Not Done",
            task.note or "",
        ])
        row = ws.max_row
        fill_color = "E5F4E0" if task.done else "FBE4E1"
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 58
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 35

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"MCQ_Checklist_{session_row.checklist_type}_{session_row.section}_{session_row.session_date.isoformat()}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )


@checklist_bp.get("/export/temperature/<int:session_id>.xlsx")
@admin_required
def export_temperature_excel(session_id: int):
    deps = _try_openpyxl()
    if deps is None:
        abort(503)
    Workbook, Font, PatternFill, Alignment, get_column_letter = deps

    session_row = TemperatureSession.query.get_or_404(session_id)
    temp_data = TEMPERATURE_RECORDS.get(session_row.record_type, {})

    wb = Workbook()
    ws = wb.active
    ws.title = "Temperature Record"

    ws["A1"] = "MCQ VIETNAMESE STREET FOOD — FOOD TEMPERATURE RECORD"
    ws["A1"].font = Font(bold=True, size=16, color="2C201B")
    ws.merge_cells("A1:M1")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = f"{temp_data.get('title', '')}   |   Date: {session_row.session_date.isoformat()}"
    ws.merge_cells("A2:M2")
    ws["A3"] = f"Recorded by: {session_row.recorded_by or ''}   |   Checked by: {session_row.checked_by or ''}"
    ws.merge_cells("A3:M3")
    ws["A4"] = "Safe zone: < 5°C or > 60°C. Danger zone 5-60°C: max 4 hours before discarding."
    ws["A4"].font = Font(italic=True, color="EB5130")
    ws.merge_cells("A4:M4")
    ws.append([""])

    headers = [
        "Food Item",
        "C1 Time", "C1 °C",
        "C2 Time", "C2 °C",
        "C3 Time", "C3 °C",
        "C4 Time", "C4 °C",
        "C5 Time", "C5 °C",
        "Discarded",
    ]
    ws.append(headers)
    hr = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=hr, column=col)
        cell.fill = PatternFill(start_color="2C201B", end_color="2C201B", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for reading in session_row.readings:
        row_data = [reading.food_name]
        for n in range(1, 6):
            row_data.append(getattr(reading, f"c{n}_time") or "")
            value = getattr(reading, f"c{n}_temp")
            row_data.append(value if value is not None else "")
        row_data.append(reading.discarded or "N")
        ws.append(row_data)
        row_num = ws.max_row
        for n in range(1, 6):
            cell = ws.cell(row=row_num, column=(n - 1) * 2 + 3)
            value = getattr(reading, f"c{n}_temp")
            if value is None:
                continue
            if value < 5 or value > 60:
                cell.fill = PatternFill(start_color="FFD2C7", end_color="FFD2C7", fill_type="solid")
                cell.font = Font(bold=True, color="EB5130")
            else:
                cell.fill = PatternFill(start_color="DDF1D1", end_color="DDF1D1", fill_type="solid")

    ws.column_dimensions["A"].width = 30
    for col in range(2, 13):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.column_dimensions[get_column_letter(12)].width = 14

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"MCQ_Temperature_{session_row.record_type}_{session_row.session_date.isoformat()}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )


@checklist_bp.get("/export/bulk.xlsx")
@admin_required
def export_bulk_excel():
    deps = _try_openpyxl()
    if deps is None:
        abort(503)
    Workbook, Font, PatternFill, Alignment, _get_column_letter = deps

    date_from = _parse_date(request.args.get("date_from"), default=date.today() - timedelta(days=30))
    date_to = _parse_date(request.args.get("date_to"))

    chk_rows = (
        ChecklistSession.query.filter(
            ChecklistSession.session_date.between(date_from, date_to)
        )
        .order_by(
            ChecklistSession.session_date.desc(),
            ChecklistSession.checklist_type,
            ChecklistSession.section,
        )
        .all()
    )
    temp_rows = (
        TemperatureSession.query.filter(
            TemperatureSession.session_date.between(date_from, date_to)
        )
        .order_by(TemperatureSession.session_date.desc(), TemperatureSession.record_type)
        .all()
    )

    wb = Workbook()
    header_fill = PatternFill(start_color="2C201B", end_color="2C201B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    ws1 = wb.active
    ws1.title = "Checklist Summary"
    ws1["A1"] = f"MCQ — Checklist Summary ({date_from.isoformat()} to {date_to.isoformat()})"
    ws1["A1"].font = Font(bold=True, size=14, color="EB5130")
    ws1.merge_cells("A1:K1")
    ws1.append([""])
    headers = [
        "Date", "Day", "Type", "Section", "Responsible", "Submitted By",
        "Tasks Done", "Total Tasks", "Completion %", "Verified", "Issues",
    ]
    ws1.append(headers)
    hr = ws1.max_row
    for col in range(1, len(headers) + 1):
        cell = ws1.cell(row=hr, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row in chk_rows:
        done, total, pct = _completion(row)
        ws1.append([
            row.session_date.isoformat(),
            row.day_of_week or "",
            CHECKLISTS.get(row.checklist_type, {}).get("title", row.checklist_type),
            row.section.title(),
            row.responsible or "",
            row.submitted_by or "",
            done,
            total,
            f"{pct}%",
            "Yes" if row.verified else "No",
            (row.overall_result or "").replace("_", " ").title(),
        ])

    ws2 = wb.create_sheet("Temperature Summary")
    ws2["A1"] = f"MCQ — Temperature Summary ({date_from.isoformat()} to {date_to.isoformat()})"
    ws2["A1"].font = Font(bold=True, size=14, color="EB5130")
    ws2.merge_cells("A1:F1")
    ws2.append([""])
    temp_headers = ["Date", "Type", "Recorded by", "Checked by", "Submitted At", "Notes"]
    ws2.append(temp_headers)
    hr = ws2.max_row
    for col in range(1, len(temp_headers) + 1):
        cell = ws2.cell(row=hr, column=col)
        cell.fill = header_fill
        cell.font = header_font
    for row in temp_rows:
        ws2.append([
            row.session_date.isoformat(),
            TEMPERATURE_RECORDS.get(row.record_type, {}).get("title", row.record_type),
            row.recorded_by or "",
            row.checked_by or "",
            row.submitted_at.strftime("%Y-%m-%d %H:%M") if row.submitted_at else "",
            row.notes or "",
        ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"MCQ_Operations_{date_from.isoformat()}_to_{date_to.isoformat()}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )
